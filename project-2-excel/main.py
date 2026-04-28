import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import os
import regex as re

# -------- EXTRACT --------
file_path = os.path.join(os.path.dirname(__file__), "messy_data.csv")
df = pd.read_csv(file_path)


# -------- TRANSFORM --------
# A. Fix casing and strip spaces
df["Name"] = df["Name"].astype(str).str.title().str.strip()
df["Email"] = df["Email"].astype(str).str.lower()
df["Department"] = df["Department"].astype(str).str.title()

# B. Drop duplicates and rows with critical missing fields
df = df.drop_duplicates()
df = df.dropna(subset=["Name", "Email"])

# C. Handle phone formatting
df["Phone"] = df["Phone"].astype(str).str.replace(r"[^0-9]", "", regex=True)

# D. Validate emails and flag bad ones
def is_valid_email(email):
    pattern = r"^[\w\.-]+@[\w\.-]+\.\w{2,}$"
    return bool(re.match(pattern, email))

df["Email_Valid"] = df["Email"].apply(lambda x: "✅" if is_valid_email(x) else "❌ Invalid")

# E. Fix impossible ages 
df.loc[df["Age"] > 100, "Age"] = None

# F. Fill missing values
df["Phone"] = df["Phone"].fillna("N/A")
df["Age"] = df["Age"].fillna("Unknown")
df["Salary"] = df["Salary"].fillna(0)


# -------- LOAD (Save to Excel) --------
output_file = "cleaned_data.xlsx"
df.to_excel(output_file, index=False, sheet_name="Cleaned Data")

# -------- STYLE THE EXCEL FILE --------
wb = openpyxl.load_workbook(output_file)
ws = wb.active

header_fill = PatternFill(start_color="2F75B6", end_color="2F75B6", fill_type="solid")
header_font =  Font(color="FFFFFF", bold=True)

for cell in ws[1]: 
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

for col in ws.columns:
    max_length = max(len(str(cell.value or "")) for cell in col)
    ws.column_dimensions[col[0].column_letter].width =  min(max_length + 4, 50)

wb.save(output_file)
print(f"\n🎉 Clean formatted Excel saved as {output_file}")

